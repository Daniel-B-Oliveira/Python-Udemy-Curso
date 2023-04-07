# openpyxl para arquivos Excel xlsx, xlsm, xltx e xltm (instalação)
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/

from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = Workbook()
# worksheet:Worksheet = workbook.active   #type: ignore

sheet_name = 'My_Spread_Sheet'
workbook.create_sheet(sheet_name)

worksheet: Worksheet = workbook[sheet_name]

workbook.remove(workbook['Sheet'])

worksheet.cell(1, 1, 'Name')
worksheet.cell(1, 2, 'Age')
worksheet.cell(1, 3, 'Grade')

students = [
    # name #age #grade
    ['João', 14, 5.5],
    ['Maria', 13, 9.7],
    ['Luiz', 15, 8.8],
    ['Alberto', 16, 10],
]

# for i, students_row in enumerate(students, start=2):
#     for j, student_collon in enumerate(students_row, start=1):
#         worksheet.cell(i, j, student_collon)

for student in students:
    worksheet.append(student)


workbook.save(WORKBOOK_PATH)
