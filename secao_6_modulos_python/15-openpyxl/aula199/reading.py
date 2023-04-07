# openpyxl - Ler dados de uma planilha
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/

from pathlib import Path

from openpyxl import Workbook, load_workbook
# from openpyxl.cell import Cell  # Problema de tipagem já resolvido
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook: Workbook = load_workbook(WORKBOOK_PATH)
# worksheet:Worksheet = workbook.active   #type: ignore

sheet_name = 'My_Spread_Sheet'
worksheet: Worksheet = workbook[sheet_name]

# row: tuple[Cell]  # Problema de tipagem já resolvido
for row in worksheet.iter_rows(min_row=2):
    for cell in row:
        print(cell.value, end='\t')

        if cell.value == 'Maria':
            worksheet.cell(cell.row, 2, value='23')
    print()

# worksheet['B3'].value = 14  # type: ignore

workbook.save(WORKBOOK_PATH)